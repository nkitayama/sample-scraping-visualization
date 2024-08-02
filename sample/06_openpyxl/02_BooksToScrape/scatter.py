from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series, marker

# Excelファイルを読み込む
workbook = load_workbook('connecting_to_data.xlsx')

# アクティブのシートを取得する
worksheet = workbook.active

# カテゴリごとのデータの数を格納する辞書を作成する
num_by_category = {}

# カテゴリごとのデータの数を取得するために、行ごとにループする
for row in worksheet.iter_rows(min_row=2, values_only=True):
    # カテゴリの値を取得する
    category = row[3]
    # 辞書のキーにカテゴリが追加されていない場合、キーを追加する
    if category not in num_by_category:
        num_by_category[category] = 0
    # カテゴリごとのデータの数をインクリメントする
    num_by_category[category] += 1

# グラフ形式を散布図に設定する
chart = ScatterChart()
# グラフのタイトルを設定する
chart.title = "Price, Star Rating"
# グラフの軸のタイトルを設定する
chart.x_axis.title = "Price"
chart.y_axis.title = "Star Rating"
# グラフのサイズを指定する
chart.width = 30
chart.height = 15

# カテゴリごとのデータの開始行の初期値を設定する
start_line = 2

# カテゴリごとにデータを追加するために、カテゴリごとにループする
for category, num in num_by_category.items():
    # x軸として扱う"price"のデータの参照範囲を指定する
    xvalues = Reference(worksheet, min_col=2, min_row=start_line, max_row=start_line+num-1)
    # y軸として扱う"Star Rating"データの参照範囲を指定する
    yvalues = Reference(worksheet, min_col=3, min_row=start_line, max_row=start_line+num-1)
    # 散布図のデータを作成する
    series = Series(yvalues, xvalues=xvalues, title=category)
    # 線を描画せず、マーカーのみを描画する
    series.graphicalProperties.line.noFill = True 
    # マーカーを円に設定する
    series.marker = marker.Marker('circle')
    # グラフに散布図のデータを追加する
    chart.series.append(series)
    # グラフのデータの開始行を更新する
    start_line += num

# シートにグラフを追加する
worksheet.add_chart(chart, "E2")

# ワークブックを保存する
workbook.save("scatter_chart.xlsx")