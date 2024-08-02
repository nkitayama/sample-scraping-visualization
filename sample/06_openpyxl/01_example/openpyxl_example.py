from openpyxl import Workbook

# Excelのワークブックに相当するインスタンスを作成する
workbook = Workbook()

# アクティブのシートを取得する
worksheet = workbook.active

# A4セルに値を入力する
worksheet['A4'] = 4
# B4セル(4行2列目のセル)に値を入力する
worksheet.cell(row=4, column=2, value=10)

# A4セルの値を取得する
cell = worksheet['A4']
print(f'A4セル: {cell.value}')
# B4セル(4行2列目のセル)の値を取得する
cell = worksheet.cell(row=4, column=2)
print(f'B4セル: {cell.value}')

# シートの一覧を取得する
print(f'シート一覧:{workbook.sheetnames}')
# 新しくシートを作成し、シート一覧を取得する
workbook.create_sheet('Mysheet')
print(f'シート一覧:{workbook.sheetnames}')

# ワークブックを保存する
workbook.save('openpyxl_example.xlsx')